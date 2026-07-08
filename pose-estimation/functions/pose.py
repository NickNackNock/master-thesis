import cv2
import numpy as np
from tqdm import tqdm
from rtmlib import draw_skeleton

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from collections import defaultdict



HEAD_KP_INDICES = [0, 1, 2, 3, 4]

PERSON_BOX_COLOR = (0, 255, 0)
HEAD_BOX_COLOR   = (255, 0, 0)

KEYPOINT_NAMES = [
    "Nose", "Left Eye", "Right Eye", "Left Ear", "Right Ear",
    "Left Shoulder", "Right Shoulder", "Left Elbow", "Right Elbow",
    "Left Wrist", "Right Wrist", "Left Hip", "Right Hip",
    "Left Knee", "Right Knee", "Left Ankle", "Right Ankle",
]


# ----- KEYPOINT DATA SAVER -----
class KeypointDataSaver:
    """Accumulates per-frame keypoint data per tracked ID, saves to .xlsx."""

    def __init__(self):
        # { track_id: [ {frame, kp_x0, kp_y0, score0, ...}, ... ] }
        self._data: dict[int, list[dict]] = defaultdict(list)

    def record(self, frame_idx: int, track_id: int,
               keypoints: np.ndarray, scores: np.ndarray):
        row = {"frame": frame_idx}

        for kp_idx, name in enumerate(KEYPOINT_NAMES):
            key = name.lower().replace(" ", "_")
            if kp_idx < len(keypoints):
                row[f"kp_{key}_x"] = float(keypoints[kp_idx, 0])
                row[f"kp_{key}_y"] = float(keypoints[kp_idx, 1])
                row[f"score_{key}"] = float(scores[kp_idx])

            else:
                row[f"kp_{key}_x"] = None
                row[f"kp_{key}_y"] = None
                row[f"score_{key}"] = None

        self._data[track_id].append(row)

    def save(self, path: str):
        wb = Workbook()
        wb.remove(wb.active)  # remove default empty sheet

        header_fill   = PatternFill("solid", start_color="4F81BD")
        header_font   = Font(bold=True, color="FFFFFF", name="Arial", size=10)

        cell_font = Font(name="Arial", size=10)

        for track_id in sorted(self._data.keys()):
            ws = wb.create_sheet(title=f"Person_{track_id}")
            rows = self._data[track_id]
            if not rows:
                continue

            headers = list(rows[0].keys())
            # Write header row
            for col, h in enumerate(headers, start=1):
                cell = ws.cell(row=1, column=col, value=h)
                cell.font   = header_font
                cell.fill   = header_fill
                cell.alignment = Alignment(horizontal="center", wrap_text=True)
                ws.column_dimensions[cell.column_letter].width = 14

            # Write data rows
            for row_idx, row_data in enumerate(rows, start=2):
                for col, h in enumerate(headers, start=1):
                    cell = ws.cell(row=row_idx, column=col, value=row_data[h])
                    cell.font = cell_font

            ws.freeze_panes = "B2"  # freeze header + frame column

        wb.save(path)
        print(f"Keypoint data saved → {path}")


# ----- BOUNDING BOXES -----
def head_bbox_from_pose(keypoints, scores, score_thresh=0.5, expansion=1.6):
    """Returns [x1, y1, x2, y2] derived from head keypoints."""
    head_kps    = keypoints[HEAD_KP_INDICES]
    head_scores = scores[HEAD_KP_INDICES]

    visible = head_scores > score_thresh
    #print(visible.sum())

    if visible.sum() <= 2:
        return None
    
    pts = head_kps[visible]
    x1, y1 = pts.min(axis=0)
    x2, y2 = pts.max(axis=0)
    cx, cy = (x1 + x2) / 2, (y1 + y2) / 2
    hw, hh = (x2 - x1) / 2 * expansion, (y2 - y1) / 2 * expansion
    return np.array([cx - hw, cy - hh, cx + hw, cy + hh])


def pose_to_bbox(keypoints: np.ndarray, expansion: float = 1.25) -> np.ndarray:
    x, y = keypoints[:, 0], keypoints[:, 1]
    bbox   = np.array([x.min(), y.min(), x.max(), y.max()])
    center = np.array([bbox[0] + bbox[2], bbox[1] + bbox[3]]) / 2
    return np.concatenate([
        center - (center - bbox[:2]) * expansion,
        center + (bbox[2:] - center) * expansion,
    ])


# ----- INPUT VIDEO SETUP -----
def setup_input_video(input_video):
    cap    = cv2.VideoCapture(input_video)
    total  = int(cap.get(cv2.CAP_PROP_FRAME_COUNT))
    fps    = cap.get(cv2.CAP_PROP_FPS)
    width  = int(cap.get(cv2.CAP_PROP_FRAME_WIDTH))
    height = int(cap.get(cv2.CAP_PROP_FRAME_HEIGHT))
    return cap, total, fps, width, height


# ----- POSE ONLY -----
def run_pose_estimation(input_video, output_video, pose_model,
                        bbox_body=None, bbox_head=None):
    cap, total, fps, width, height = setup_input_video(input_video)
    writer = cv2.VideoWriter(output_video, cv2.VideoWriter_fourcc(*'mp4v'),
                             fps, (width, height))
    with tqdm(total=total, unit='frame', desc='Processing') as pbar:
        while True:
            ret, img = cap.read()
            if not ret:
                break
            keypoints, scores = pose_model(img)
            img = draw_skeleton(img, keypoints, scores, kpt_thr=0.5)
            for person_kps, person_scores in zip(keypoints, scores):
                if bbox_body:
                    bbox = pose_to_bbox(person_kps[:, :2])
                    x1, y1, x2, y2 = bbox.astype(int)
                    cv2.rectangle(img, (x1, y1), (x2, y2), PERSON_BOX_COLOR, 2)
                if bbox_head:
                    head_bbox = head_bbox_from_pose(person_kps[:, :2], person_scores)
                    if head_bbox is not None:
                        hx1, hy1, hx2, hy2 = head_bbox.astype(int)
                        cv2.rectangle(img, (hx1, hy1), (hx2, hy2), HEAD_BOX_COLOR, 2)
            writer.write(img)
            pbar.update(1)
    cap.release()
    writer.release()
    print(f'Saved to {output_video}')


# ----- POSE + TRACKING (with data export) -----
def run_pose_estimation_with_tracking(input_video, output_video, pose_model,
                                      tracker, save_xlsx: str | None = None):
    cap, total, fps, width, height = setup_input_video(input_video)
    writer = cv2.VideoWriter(output_video, cv2.VideoWriter_fourcc(*'mp4v'),
                             fps, (width, height))

    saver = KeypointDataSaver() if save_xlsx else None
    frame_idx = 0

    with tqdm(total=total, unit="frame", desc="Tracking") as pbar:
        while True:
            ret, frame = cap.read()
            if not ret:
                break
            
            """keypoints and scores hold one entry per person detected by
                the pose model this frame
                No identity, no memory of previous frames, just raw detections
                in whatever order the model spit them out.
            """
            keypoints, scores = pose_model(frame)
            frame = draw_skeleton(frame, keypoints, scores, kpt_thr=0.5)

            """" Collapses of 17 keypoints into a single bounding box for the tracker"""
            dets_list, pose_mapping = [], {}
            for i, (p_kps, p_scores) in enumerate(zip(keypoints, scores)):
                bbox = pose_to_bbox(p_kps[:, :2])
                dets_list.append([*bbox, float(p_scores.mean()), 0.0])
                pose_mapping[len(dets_list) - 1] = (p_kps, p_scores)

            """ Feeds the current frame's detections to the tracker, 
                which assigns them IDs based on its internal logic (e.g. motion, appearance)
            """
            dets   = np.array(dets_list) if dets_list else np.empty((0, 6))
            tracks = tracker.update(dets, frame)

            if tracks is not None and len(tracks):
                for track in tracks:
                    x1, y1, x2, y2, track_id, conf, cls, ind = track
                    x1, y1, x2, y2 = int(x1), int(y1), int(x2), int(y2)
                    track_id, ind   = int(track_id), int(ind)

                    cv2.rectangle(frame, (x1, y1), (x2, y2), PERSON_BOX_COLOR, 1)
                    cv2.putText(frame, f"ID {track_id}", (x1, y1 - 10),
                                cv2.FONT_HERSHEY_SIMPLEX, 0.6, PERSON_BOX_COLOR, 2)

                    if ind in pose_mapping:
                        p_kps, p_scores = pose_mapping[ind]

                        # Record data
                        if saver:
                            saver.record(frame_idx, track_id, p_kps[:, :2], p_scores)

                        head_bbox = head_bbox_from_pose(p_kps[:, :2], p_scores)
                        if head_bbox is not None:
                            hx1, hy1, hx2, hy2 = head_bbox.astype(int)
                            cv2.rectangle(frame, (hx1, hy1), (hx2, hy2), HEAD_BOX_COLOR, 2)
                            cv2.putText(frame, f"Head {track_id}", (hx1, hy1 - 6),
                                        cv2.FONT_HERSHEY_SIMPLEX, 0.45, HEAD_BOX_COLOR, 1)

            writer.write(frame)
            pbar.update(1)
            frame_idx += 1

    cap.release()
    writer.release()
    print(f'Video saved to {output_video}')

    if saver and save_xlsx:
        saver.save(save_xlsx)
