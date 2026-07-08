"""
Stereo 3D pose reconstruction.

Prerequisites (run these first via main.py):
  - Option 1 → run on INPUT_VIDEO       → OUTPUT_RAW_POSE_XLSX
  - Option 1 → run on INPUT_VIDEO_CAM2  → OUTPUT_RAW_POSE_XLSX_CAM2
  - Option 2 → filter cam1              → OUTPUT_FILT_POSE_XLSX
  - Option 2 → filter cam2              → OUTPUT_FILT_POSE_XLSX_CAM2

Then run Option 4 in main.py which calls reconstruct_3d() here.

Pipeline per frame
──────────────────
  1. Load filtered 2D keypoints from both cameras (already smoothed).
  2. Match persons across cameras via epipolar geometry (Sampson distance +
     Hungarian algorithm).  This is necessary because BoxMOT assigns IDs
     independently in each camera — there is no guarantee that ID 1 in cam1
     is the same person as ID 1 in cam2.
  3. Undistort matched keypoints using each camera's intrinsics.
  4. Triangulate → metric 3D points in cam1's coordinate frame.
  5. Save per-person 3D keypoints to xlsx (one sheet per person, columns
     kp_<name>_x/y/z plus the averaged confidence score).

Coordinate system of the output
────────────────────────────────
  The 3D points are expressed in cam1's coordinate frame:
    - Origin at cam1's optical centre
    - Z points along cam1's optical axis (away from the camera)
    - X right, Y down (standard OpenCV convention)
  All metric units match the unit used during stereo calibration (usually mm).
"""

import numpy as np
import pandas as pd
import cv2
from pathlib import Path
from scipy.optimize import linear_sum_assignment
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from collections import defaultdict
from tqdm import tqdm

KEYPOINT_NAMES = [
    "Nose", "Left Eye", "Right Eye", "Left Ear", "Right Ear",
    "Left Shoulder", "Right Shoulder", "Left Elbow", "Right Elbow",
    "Left Wrist", "Right Wrist", "Left Hip", "Right Hip",
    "Left Knee", "Right Knee", "Left Ankle", "Right Ankle",
]


def _kp_col(name: str) -> str:
    """'Left Eye' → 'left_eye'  (must match KeypointDataSaver convention in pose.py)"""
    return name.lower().replace(" ", "_")


# ---------------------------------------------------------------------------
# Stereo calibration loader
# ---------------------------------------------------------------------------

def load_stereo_calibration(calib_path: str) -> dict:
    """
    Load stereo calibration from a .npz file — the standard output of
    cv2.stereoCalibrate() saved with np.savez().

    Required keys in the file: K1, d1, K2, d2, R, T
      K1, K2  : 3x3 camera intrinsic matrices
      d1, d2  : distortion coefficient vectors (4, 5, 8, 12, or 14 elements)
      R       : 3x3 rotation matrix from cam1 to cam2
      T       : 3x1 translation vector from cam1 to cam2

    Derived quantities computed here:
      E  : essential matrix  (E = [T]x R)
      F  : fundamental matrix (F = K2^{-T} E K1^{-1})
      P1 : 3x4 projection matrix for cam1  (K1 [I | 0])
      P2 : 3x4 projection matrix for cam2  (K2 [R | T])

    Note: if you have already run cv2.stereoRectify() and saved R1/R2/P1/P2,
    pass those projection matrices instead of the ones computed here and
    use cv2.undistortPoints with the rectification rotation R1/R2.
    """
    base_dir = Path(calib_path)

    # Load each matrix using numpy's text loader
    try:
        K1 = np.loadtxt(base_dir / "camera_matrix_cam1.txt")
        d1 = np.loadtxt(base_dir / "distortion_coefficient_cam1.txt")
        K2 = np.loadtxt(base_dir / "camera_matrix_cam2.txt")
        d2 = np.loadtxt(base_dir / "distortion_coefficient_cam2.txt")
        R  = np.loadtxt(base_dir / "Rotation_matrix.txt")
        T  = np.loadtxt(base_dir / "Translation_Matrix.txt").flatten()  # ensure shape (3,)
    except FileNotFoundError as e:
        raise FileNotFoundError(
            f"Could not find all required calibration files in '{calib_path}'. "
            f"Ensure files match expected names exactly. Error details: {e}"
        )

    # Essential matrix: E = [T]× R
    Tx = np.array([[ 0,    -T[2],  T[1]],
                   [ T[2],  0,    -T[0]],
                   [-T[1],  T[0],  0   ]])
    E = Tx @ R

    # Fundamental matrix: F = K2^{-T} E K1^{-1}
    F = np.linalg.inv(K2).T @ E @ np.linalg.inv(K1)

    # Projection matrices (cam1 at world origin)
    P1 = K1 @ np.hstack([np.eye(3),    np.zeros((3, 1))])
    P2 = K2 @ np.hstack([R,            T.reshape(3, 1) ])

    calib.update({"E": E, "F": F, "P1": P1, "P2": P2, "T": T})
    return calib


# ---------------------------------------------------------------------------
# Keypoint I/O
# ---------------------------------------------------------------------------

def load_pose_xlsx_for_stereo(xlsx_path: str, frame_offset: int = 0,
                               kp_names=KEYPOINT_NAMES) -> dict:
    """
    Read a filtered (or raw) pose xlsx and return a nested dict:
        { frame_idx : { track_id : (kps (17,2), scores (17,)) } }

    frame_offset shifts all frame indices by a constant — use this to
    compensate for cameras that were not started at exactly the same time.
    Positive offset means this camera's frames are shifted forward
    (i.e. frame 0 here corresponds to frame `offset` in the other camera).
    """
    sheets = pd.read_excel(xlsx_path, sheet_name=None)
    frame_data: dict[int, dict] = defaultdict(dict)

    for sheet_name, df in sheets.items():
        # Sheet names follow the "Person_<id>" convention (KeypointDataSaver in pose.py)
        track_id = int(sheet_name.split("_")[1])

        for _, row in df.iterrows():
            frame_idx = int(row["frame"]) + frame_offset

            kps = np.array(
                [[row[f"kp_{_kp_col(n)}_x"], row[f"kp_{_kp_col(n)}_y"]]
                 for n in kp_names],
                dtype=float,
            )
            scores = np.array(
                [row[f"score_{_kp_col(n)}"] for n in kp_names],
                dtype=float,
            )
            frame_data[frame_idx][track_id] = (kps, scores)

    print(f"  {len(sheets)} person sheet(s), "
          f"{sum(len(v) for v in frame_data.values())} total frame-detections "
          f"(offset={frame_offset:+d})")
    return dict(frame_data)


# ---------------------------------------------------------------------------
# Keypoint undistortion
# ---------------------------------------------------------------------------

def undistort_keypoints(kps: np.ndarray, K: np.ndarray,
                         d: np.ndarray) -> np.ndarray:
    """
    Remove lens distortion from 2D keypoints.

    This step is mandatory before triangulation — skipping it introduces
    systematic errors in 3D position that grow toward the image periphery.

    kps : (17, 2) array in pixel coordinates (may contain zeros for missing joints)
    K   : 3×3 intrinsic matrix
    d   : distortion coefficients

    Returns (17, 2) undistorted pixel coordinates.
    The P=K argument keeps the output in the original (unrectified) pixel space,
    which matches the projection matrices P1/P2 we built in load_stereo_calibration.
    """
    # Only process non-zero (actually detected) keypoints
    valid = (kps[:, 0] != 0) | (kps[:, 1] != 0)
    result = kps.copy()

    if valid.sum() > 0:
        pts = kps[valid].reshape(-1, 1, 2).astype(np.float64)
        undistorted = cv2.undistortPoints(pts, K, d, P=K)
        result[valid] = undistorted.reshape(-1, 2)

    return result


# ---------------------------------------------------------------------------
# Cross-camera person matching
# ---------------------------------------------------------------------------

def _sampson_distance(kps1: np.ndarray, kps2: np.ndarray,
                      F: np.ndarray) -> float:
    """
    Mean Sampson distance between keypoint pairs under the fundamental matrix F.

    The Sampson distance is a first-order approximation to the reprojection
    error under the epipolar constraint x2^T F x1 = 0.  It is symmetric and
    more numerically stable than the plain algebraic error.

    Only keypoints present in both views (non-zero in both arrays) are used.
    Returns np.inf when fewer than 2 such pairs exist (no reliable estimate).
    """
    valid = ((kps1[:, 0] != 0) | (kps1[:, 1] != 0)) & \
            ((kps2[:, 0] != 0) | (kps2[:, 1] != 0))

    if valid.sum() < 2:
        return np.inf

    p1 = np.hstack([kps1[valid], np.ones((valid.sum(), 1))])   # (N, 3)
    p2 = np.hstack([kps2[valid], np.ones((valid.sum(), 1))])   # (N, 3)

    Fp1  = (F   @ p1.T).T   # epipolar lines in cam2  (N, 3)
    FTp2 = (F.T @ p2.T).T  # epipolar lines in cam1  (N, 3)

    num  = np.sum(p2 * Fp1, axis=1) ** 2
    den  = Fp1[:, 0]**2 + Fp1[:, 1]**2 + FTp2[:, 0]**2 + FTp2[:, 1]**2
    dists = num / (den + 1e-10)

    return float(np.mean(dists))


def match_persons(persons_cam1: dict, persons_cam2: dict,
                  F: np.ndarray) -> list[tuple[int, int]]:
    """
    Match people between the two cameras for a single frame.

    Builds a cost matrix of Sampson distances (N_persons_cam1 × N_persons_cam2)
    and solves the optimal assignment with the Hungarian algorithm.

    Returns a list of (track_id_cam1, track_id_cam2) pairs.
    Unmatched persons (e.g. someone visible in only one camera) are excluded.

    Why not just match by track_id equality?
    ─────────────────────────────────────────
    BoxMOT assigns IDs independently in each camera view.  There is no
    guarantee that "Person_1" in cam1 is the same physical person as
    "Person_1" in cam2 — the ID assignment depends on detection order and
    the tracker's internal state, which differ between views.
    """
    ids1 = list(persons_cam1.keys())
    ids2 = list(persons_cam2.keys())

    if not ids1 or not ids2:
        return []

    # Build pairwise Sampson distance cost matrix
    cost = np.full((len(ids1), len(ids2)), np.inf)
    for i, id1 in enumerate(ids1):
        kps1, _ = persons_cam1[id1]
        for j, id2 in enumerate(ids2):
            kps2, _ = persons_cam2[id2]
            cost[i, j] = _sampson_distance(kps1, kps2, F)

    # Hungarian algorithm needs finite values — replace inf with a large number
    cost_finite = np.where(np.isinf(cost), 1e9, cost)
    row_ind, col_ind = linear_sum_assignment(cost_finite)

    # Only accept matches where the Sampson distance was finite (i.e. enough
    # overlapping keypoints existed to form a meaningful estimate)
    matches = [
        (ids1[r], ids2[c])
        for r, c in zip(row_ind, col_ind)
        if cost[r, c] < 1e8
    ]
    return matches


# ---------------------------------------------------------------------------
# Triangulation
# ---------------------------------------------------------------------------

def triangulate_keypoints(kps1: np.ndarray, kps2: np.ndarray,
                           P1: np.ndarray, P2: np.ndarray) -> np.ndarray:
    """
    Triangulate matched 2D keypoints from two calibrated cameras.

    kps1, kps2 : (17, 2) — undistorted pixel coordinates from each camera
    P1, P2     : (3, 4)  — projection matrices

    Returns (17, 3) 3D keypoints in cam1's coordinate frame (metric units).
    Keypoints missing in either view (zero in either kps array) are returned
    as (0, 0, 0) so that downstream code using the same zero-sentinel convention
    can filter them out.

    cv2.triangulatePoints is numerically robust but uses a simple DLT;
    for higher accuracy with noisy data you could refine using
    cv2.correctMatches() (optimal triangulation) before calling this function.
    """
    n = len(kps1)
    kps3d = np.zeros((n, 3))

    # Only triangulate keypoints that are present in BOTH views
    valid = ((kps1[:, 0] != 0) | (kps1[:, 1] != 0)) & \
            ((kps2[:, 0] != 0) | (kps2[:, 1] != 0))

    if valid.sum() == 0:
        return kps3d

    pts1 = kps1[valid].T.astype(np.float64)   # (2, N)
    pts2 = kps2[valid].T.astype(np.float64)   # (2, N)

    # triangulatePoints returns homogeneous coordinates (4, N)
    pts4d = cv2.triangulatePoints(P1, P2, pts1, pts2)

    # Convert from homogeneous to Euclidean by dividing by w
    pts3d = (pts4d[:3] / pts4d[3]).T          # (N, 3)
    kps3d[valid] = pts3d

    return kps3d


# ---------------------------------------------------------------------------
# Output
# ---------------------------------------------------------------------------

def _save_3d_xlsx(person_3d: dict, path: str, kp_names=KEYPOINT_NAMES):
    """
    Save 3D keypoints to xlsx.  One sheet per person (named Person_<id>_3D).
    Columns: frame, then kp_<name>_x/y/z and score_<name> for each keypoint.
    Headers are red to visually distinguish from the 2D xlsx files.
    """
    wb = Workbook()
    wb.remove(wb.active)

    # Red header to visually distinguish 3D sheets from the blue 2D sheets
    header_fill = PatternFill("solid", start_color="C0504D")
    header_font = Font(bold=True, color="FFFFFF", name="Arial", size=10)
    cell_font   = Font(name="Arial", size=10)

    for track_id in sorted(person_3d.keys()):
        ws   = wb.create_sheet(title=f"Person_{track_id}_3D")
        rows = person_3d[track_id]
        if not rows:
            continue

        headers = list(rows[0].keys())
        for col, h in enumerate(headers, start=1):
            cell           = ws.cell(row=1, column=col, value=h)
            cell.font      = header_font
            cell.fill      = header_fill
            cell.alignment = Alignment(horizontal="center", wrap_text=True)
            ws.column_dimensions[cell.column_letter].width = 14

        for row_idx, row_data in enumerate(rows, start=2):
            for col, h in enumerate(headers, start=1):
                cell      = ws.cell(row=row_idx, column=col, value=row_data[h])
                cell.font = cell_font

        ws.freeze_panes = "B2"

    wb.save(path)
    print(f"3D keypoints saved → {path}")


# ---------------------------------------------------------------------------
# Main entry point
# ---------------------------------------------------------------------------

def reconstruct_3d(xlsx_cam1: str, xlsx_cam2: str, calib_path: str,
                   out_path: str, frame_offset: int = 0,
                   kp_names=KEYPOINT_NAMES) -> dict:
    """
    Full stereo 3D reconstruction pipeline.

    Parameters
    ──────────
    xlsx_cam1     : filtered pose xlsx for camera 1 (OUTPUT_FILT_POSE_XLSX)
    xlsx_cam2     : filtered pose xlsx for camera 2 (OUTPUT_FILT_POSE_XLSX_CAM2)
    calib_path    : path to stereo calibration .npz file (STEREO_CALIB_FILE)
    out_path      : where to write the 3D xlsx (OUTPUT_3D_XLSX)
    frame_offset  : cam2 frame index = cam1 frame index + frame_offset
                    (0 if cameras were hardware-synced)
    kp_names      : keypoint name list (must match what pose.py used)

    Returns
    ───────
    person_3d : { cam1_track_id : [ {frame, kp_*_x/y/z, score_*}, ... ] }
                The 3D keypoints in cam1's coordinate frame.

    Typical call from main.py (option 4):
        from stereo_reconstruction import reconstruct_3d
        reconstruct_3d(
            xlsx_cam1    = str(OUTPUT_FILT_POSE_XLSX),
            xlsx_cam2    = str(OUTPUT_FILT_POSE_XLSX_CAM2),
            calib_path   = STEREO_CALIB_FILE,
            out_path     = str(OUTPUT_3D_XLSX),
            frame_offset = STEREO_FRAME_OFFSET,
        )
    """
    print("Loading stereo calibration...")
    calib = load_stereo_calibration(calib_path)
    K1, d1 = calib["K1"], calib["d1"]
    K2, d2 = calib["K2"], calib["d2"]
    P1, P2 = calib["P1"], calib["P2"]
    F      = calib["F"]

    print(f"Loading cam1 pose data from:\n  {xlsx_cam1}")
    data1 = load_pose_xlsx_for_stereo(xlsx_cam1, frame_offset=0, kp_names=kp_names)

    print(f"Loading cam2 pose data from:\n  {xlsx_cam2}")
    data2 = load_pose_xlsx_for_stereo(xlsx_cam2, frame_offset=frame_offset,
                                       kp_names=kp_names)

    # Process only frames where both cameras have at least one detection
    all_frames = sorted(set(data1.keys()) & set(data2.keys()))
    print(f"\nFrames with detections in both cameras: {len(all_frames)}")

    person_3d: dict[int, list] = defaultdict(list)

    # Track how often each cam1↔cam2 ID pair is matched — useful for
    # diagnosing ID swaps (e.g. if the same pair doesn't dominate, the
    # matcher might be unreliable for that person)
    match_counts: dict[tuple, int] = defaultdict(int)

    for frame_idx in tqdm(all_frames, unit="frame", desc="Triangulating"):
        persons1 = data1.get(frame_idx, {})
        persons2 = data2.get(frame_idx, {})

        if not persons1 or not persons2:
            continue

        matches = match_persons(persons1, persons2, F)

        for id1, id2 in matches:
            match_counts[(id1, id2)] += 1

            kps1_raw, scores1 = persons1[id1]
            kps2_raw, scores2 = persons2[id2]

            # Undistort before triangulation — mandatory step
            kps1_u = undistort_keypoints(kps1_raw, K1, d1)
            kps2_u = undistort_keypoints(kps2_raw, K2, d2)

            kps3d = triangulate_keypoints(kps1_u, kps2_u, P1, P2)

            # Average confidence from both views as the 3D score
            combined_scores = (scores1 + scores2) / 2.0

            row: dict = {"frame": frame_idx}
            for i, name in enumerate(kp_names):
                key = _kp_col(name)
                row[f"kp_{key}_x"] = float(kps3d[i, 0])
                row[f"kp_{key}_y"] = float(kps3d[i, 1])
                row[f"kp_{key}_z"] = float(kps3d[i, 2])
                row[f"score_{key}"] = float(combined_scores[i])

            person_3d[id1].append(row)

    # Print match statistics — useful for sanity-checking the cross-camera ID assignment
    print("\nCross-camera ID match counts (cam1_id → cam2_id : n_frames):")
    for (id1, id2), count in sorted(match_counts.items()):
        print(f"  Person_{id1} (cam1) ↔ Person_{id2} (cam2) : {count} frames")
    print()

    _save_3d_xlsx(dict(person_3d), out_path, kp_names)
    return dict(person_3d)
